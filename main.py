import fdb
import configparser
import traceback
import datetime
from openpyxl.workbook import Workbook
import os
ConfigName = "CONFIG.ini"

    
def extractData(idClifor,cursor):
    safras = getSafras(cursor)
    especies = getSpecies(cursor)
    date = datetime.datetime.now().strftime("%d.%m.%Y")
    command = """
    select
    first 1
    safra.id_safra,
    safra.descricao safra,
    arm_especie.descricao especie,
    ARM_GET_SALCLIFORFS_ATU.rsaldo_seco,
    (ARM_GET_SALCLIFORFS_ATU.rsaldo_seco / coalesce(nullif(arm_especie.peso_saca,0), 60)) saldo_seco_sc,
    ARM_GET_SALCLIFORFS_ATU.rsaldo_verde,
    (ARM_GET_SALCLIFORFS_ATU.rsaldo_verde / coalesce(nullif(arm_especie.peso_saca,0), 60)) saldo_verde_sc
    from ARM_GET_SALCLIFORFS_ATU(10, {pid_safra} , {pid_especie}, {pid_clifor}, {data})
    join safra
    on (safra.id_safra = {pid_safra})
    join arm_especie
    on (arm_especie.id_especie = {pid_especie})
    join clifor
    on (clifor.id = {pid_clifor})

    """ 
    # ordem dos parametros: safra, especie, clifor , data
    resultado = []
    totalizadores = []
    index = 0
    # even safra and especie ia list of tulples so we must get the fist of each iterable element
    for especie in especies:
        print("Especie: ",especie[1])
        index = index + 1
        totalVerde = totalSeco = secoSaca = verdeSaca = auxiliar = 0
        
        for safra in safras:
            
            cursor.execute(command.format(pid_safra=safra[0],pid_especie=especie[0],pid_clifor =idClifor,data=f"'{date}'"))
              
            aux = cursor.fetchall()
            
            
            if(aux != []):
                
                resultado.append(*aux) 
                
                totalSeco = totalSeco + aux[0][3]
                secoSaca = secoSaca + aux[0][4]            
                totalVerde = totalVerde + aux[0][5]
                verdeSaca = verdeSaca + aux[0][6] 
                        
        auxiliar = (especie[0],especie[1],totalSeco,secoSaca,totalVerde,verdeSaca)       
        totalizadores.append(auxiliar)  
        #print(totalizadores) 
                
    
    #print(totalizadores)
    #print(resultado)
 
    return resultado, totalizadores

def creatExcel(Clifor,datas,totalizador, safras):
    date = datetime.datetime.now().strftime("%d_%m_%Y")
    row = 0
    column = 0
    wb = Workbook()
    
    work = wb.active
    work.title = Clifor    
    
    # the initial position of the name of the clifor will be in the 
    # B2 e C2 merged cell
    
    row = column = 2
    
    work.merge_cells(start_row=row,start_column=2,end_row=row,end_column =column +1)
    work.cell(row=row,column=column,value=Clifor)
    
    row = row + 2
    
    # each "safra" will be a tuple, bein the first positional argument the ID of the safra 
    # and the second posiotion argument the description of the safra
    for safra in safras:
        work.merge_cells(start_row=row,start_column=2,end_row=row,end_column =column +4)
        work.cell(row=row,column=column,value=safra[1])
        row = row + 1
        work.cell(row=row,column=column ,value = "Especie")
        work.cell(row=row,column=column + 1,value = "Saldo Seco Kg")
        work.cell(row=row,column=column + 2,value = "Saldo Seco SC")
        work.cell(row=row,column=column + 3,value = "Sado Verde Kg")
        work.cell(row=row,column=column + 4,value = "Saldo Verde Sc")

        row = row + 1
        
        for data in datas:
            if (data[0] == safra[0]):
                #dadoInscrever = (data[1],int(data[2]))
                work.cell(row=row,column=column,value=data[2])
                work.cell(row=row,column=column+1,value=f"{int(data[3])}")
                work.cell(row=row,column=column+2,value=f"{int(data[4])}")
                work.cell(row=row,column=column+3,value=f"{int(data[5])}")
                work.cell(row=row,column=column+4,value=f"{int(data[6])}")

                row = row + 1
                
        row = row + 1 
    
    # puttin the totalizador
    
    # will always start in the row 4 and column = 8
    column = 9
    row = 4
    
    work.merge_cells(start_row=row,start_column=column,end_row=row,end_column =column +4)
    work.cell(row=row,column=column,value="Total por especie")  
    row = row + 1
    
    work.cell(row=row,column=column ,value = "Especie")
    work.cell(row=row,column=column + 1,value = "Saldo Seco Kg")
    work.cell(row=row,column=column + 2,value = "Saldo Seco SC")
    work.cell(row=row,column=column + 3,value = "Sado Verde Kg")
    work.cell(row=row,column=column + 4,value = "Saldo Verde Sc") 
    row = row + 1
    
    for total in totalizador:
        work.cell(row=row,column=column,value=total[1])
        work.cell(row=row,column=column+1,value=f"{int(total[2])}")
        work.cell(row=row,column=column+2,value=f"{int(total[3])}")
        work.cell(row=row,column=column+3,value=f"{int(total[4])}")
        work.cell(row=row,column=column+4,value=f"{int(total[5])}")
        row = row + 1
        
    savePath = getSavePath()
    path = f"{savePath}\\saldo_{date}.xlsx"
    wb.save(path)
    
    return path
        
def getSpecies(cursor):
    command = """
    select id_especie,descricao
    from arm_especie
    order by descricao
    """
    
    try:
        cursor.execute(command)
    except:
        raise Exception("Erro ao executar comando GetEspecies")
    else: 
        especies = []
        for especie in cursor:
            especies.append(especie)
        return especies
    
def getSafras(cursor):
    command = """
        select id_safra , descricao from safra
    """   
    try:
        cursor.execute(command)
    except:
        raise Exception("Erro ao executar comando GetSafra")
    else: 
        safras = []
        for safra in cursor:
            safras.append(safra)
        return safras
 
# Read the INI file and get the path to the database   
def getPath():
    config = configparser.ConfigParser()
    config.sections() 
    config.read(ConfigName)
    
    DataBasePath = config["DATAPTH"]["DatabasePath"]
    return DataBasePath

def getSavePath():
    config = configparser.ConfigParser()
    config.sections() 
    config.read(ConfigName)
    
    DataSavePath = config["DATAPTH"]["SavePath"]
    return DataSavePath

def searchClifor(id_clifor,con):
    command = """
    select nome from clifor
    where id ="""
    
    con.execute(f"{command}{id_clifor}")
    
    # the return of fetchall is a list of tuples so we must get the position  
    name = con.fetchall()
    
    # returned a empty name we filter the error
    try:
        name = name[0][0]
    except IndexError:
       raise IndexError("Id do Clifor não encontrado")
    except Exception:
        print("to aqui (0)")
        raise Exception("Erro nao catalogado")
    else:
        return name
    
def main():
    dataBasePath = getPath()
    print("Caminho Base: ",dataBasePath)
    try:
        connection = fdb.connect(dsn=dataBasePath,password="masterkey",user="SYSDBA")
    except:
        print("Erro ao conectar ao parceiro")
    else:
        cursor = connection.cursor()
        safras = getSafras(cursor)
        
        
        print("Saldo Total de um produtor")
        
        while(1):
            
            id_clifor = int(input("Digite o ID do clifor:"))
            
            try:
                
                cliforName = searchClifor(id_clifor,cursor)
            except IndexError as value:
                print(traceback.format_exception_only(type(value),value)[0])
                continue
            except Exception as error:
                print(traceback.format_exception_only(type(error),error[0]))
            else:
                print(f"Criar relatorio para o {cliforName}")
                opcao = int(input("1- Sim    2- Não  3- sair \n Digite a opcao: "))
                
                if opcao == 1:
                    print("Extraindo ",end= " ")
                    report,totalizador = extractData(id_clifor, cursor)
                    #print("reporte ------------------ \n", report)
                    #print("Totalizador ---------------- \n",totalizador)
                    path = creatExcel(Clifor=cliforName,datas=report,totalizador=totalizador,safras=safras)
                    os.startfile(path)
                    connection.close()
                    break
                elif opcao == 3:
                    connection.close()
                    exit()
        
if "__main__" == __name__:
    print(__name__)
    main()

print("Estou aqui (0)", __name__)