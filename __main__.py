
from  PyPDF2 import PdfFileWriter
import fdb
import configparser
import traceback
import datetime
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Alignment

ConfigName = "CONFIG.ini"

    
def extractData(idClifor,cursor):
    safras = getSafras(cursor)
    especies = getSpecies(cursor)
    date = datetime.datetime.now().strftime("%d.%m.%Y")
    command = """
    select
    first 1
    safra.id_safra,
    safra.descricao safra,
    arm_especie.descricao especie,
    ARM_GET_SALCLIFORFS_ATU.rsaldo_verde saldo_verde_kg,
    (ARM_GET_SALCLIFORFS_ATU.rsaldo_verde / coalesce(nullif(arm_especie.peso_saca,0), 60)) saldo_verde_sc,
    ARM_GET_SALCLIFORFS_ATU.rsaldo_seco saldo_seco_kg,
    (ARM_GET_SALCLIFORFS_ATU.rsaldo_seco / coalesce(nullif(arm_especie.peso_saca,0), 60)) saldo_seco_sc
    from ARM_GET_SALCLIFORFS_ATU(10, {pid_safra} , {pid_especie}, {pid_clifor}, {data})
    join safra
    on (safra.id_safra = {pid_safra})
    join arm_especie
    on (arm_especie.id_especie = {pid_especie})
    join clifor
    on (clifor.id = {pid_clifor})

    """ 
    # ordem dos parametros: safra, especie, clifor , data
    resultado = []
    #print("safras",safras)
    #print("especies", especies)
    #print("data",date)
    
    # even safra and especie ia list of tulples so we must get the fist of each iterable element
    for safra in safras:
        
        for especie in especies:
            # print("comando: ",safra[0],especie[0],idClifor,f"'{date}'")
            print(command.format(pid_safra=safra[0],pid_especie=especie[0],pid_clifor =idClifor,data=f"'{date}'"))
            #cursor.execute(command%(safra[0],especie[0],idClifor,date))
            
            cursor.execute(command.format(pid_safra=safra[0],pid_especie=especie[0],pid_clifor =idClifor,data=f"'{date}'"))

            aux = cursor.fetchall()
            
            if(aux != []):
                print(aux)
                resultado.append(aux) 
    
    print(resultado)
 
    return resultado

def creatExcel(Clifor,datas,cursor):
    fields = [Clifor,"Safra","Insumo", "Saldo"]
    safras = getSafras(cursor)
    date = datetime.datetime.now().strftime("%X").replace(':','_')
    row = 0
    column = 0
    wb = Workbook()
    
    work = wb.active
    work.title = Clifor
    work.auto_filter.ref = "B4"
    
    # the initial position of the name of the clifor will be in the 
    # B2 e C2 merged cell
    
    row = column = 2
    
    work.merge_cells(start_row=row,start_column=2,end_row=row,end_column =column +1)
    work.cell(row=row,column=column,value=Clifor)
    
    row = row + 3
    
    # each "safra" will be a tuple, bein the first positional argument the ID of the safra 
    # and the second posiotion argument the description of the safra
    for safra in safras:
        
        work.merge_cells(start_row=row,start_column=2,end_row=row,end_column =column +1)
        work.cell(row=row,column=column,value=safra[1])
        row = row + 1
        
        for data in datas:
            if (data[3] == safra[0]):
                #dadoInscrever = (data[1],int(data[2]))
                work.cell(row=row,column=column,value=data[1])
                work.cell(row=row,column=column+1,value=f"{data[2]} Kg")
                row = row + 1
                
        row = row + 1 
    savePath = getSavePath()
    wb.save(f"{savePath}\\saldo_{date}.xlsx")
        
def getSpecies(cursor):
    command = """
    select id_especie,descricao
    from arm_especie
    """
    
    try:
        cursor.execute(command)
    except:
        raise Exception("Erro ao executar comando GetEspecies")
    else: 
        especies = []
        for especie in cursor:
            especies.append(especie)
        return especies
    
def getSafras(cursor):
    command = """
        select id_safra , descricao from safra
    """   
    try:
        cursor.execute(command)
    except:
        raise Exception("Erro ao executar comando GetSafra")
    else: 
        safras = []
        for safra in cursor:
            safras.append(safra)
        return safras
 
# Read the INI file and get the path to the database   
def getPath():
    config = configparser.ConfigParser()
    config.sections() 
    config.read(ConfigName)
    
    DataBasePath = config["DATAPTH"]["DatabasePath"]
    return DataBasePath

def getSavePath():
    config = configparser.ConfigParser()
    config.sections() 
    config.read(ConfigName)
    
    DataSavePath = config["DATAPTH"]["SavePath"]
    return DataSavePath

def searchClifor(id_clifor,con):
    command = """
    select nome from clifor
    where id ="""
    
    con.execute(f"{command}{id_clifor}")
    
    # the return of fetchall is a list of tuples so we must get the position  
    name = con.fetchall()
    
    # returned a empty name we filter the error
    try:
        name = name[0][0]
    except IndexError:
       raise IndexError("Id do Clifor não encontrado")
    except Exception:
        raise Exception("Erro nao catalogado")
    else:
        return name
    
def main():
    dataBasePath = getPath()
    connection = fdb.connect(dsn= f"{dataBasePath}",password="masterkey",user="SYSDBA")
    cursor = connection.cursor()
    
    print(dataBasePath)
    print("Saldo Total de um produtor")
    
    while(1):
        
        id_clifor = int(input("Digite o ID do clifor:"))
        
        try:
            cliforName = searchClifor(id_clifor,cursor)
        except IndexError as value:
            print(traceback.format_exception_only(type(value),value)[0])
            continue
        except Exception as error:
            print(traceback.format_exception_only(type(error),error))
           
        else:
            print(f"Criar relatorio para o {cliforName}")
            opcao = int(input("1- Sim    2- Não  3- sair \n Digite a opcao: "))
            
            if opcao == 1:
                report = extractData(id_clifor, cursor)
                print(report)
                creatExcel(cliforName,report,cursor)
                connection.close()
                break
            elif opcao == 3:
                connection.close()
                exit()
        
if "__main__" == __name__:
    main()
